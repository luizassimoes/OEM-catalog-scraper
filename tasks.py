import re
import time
import json
import logging
import urllib.parse
from datetime import datetime, timedelta
from dateutil import parser
from dateutil.relativedelta import relativedelta
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

class OEMCatalogScraper:

    def __init__(self):
        self.driver = None
        self.logger = logging.getLogger(__name__)
        self.wait = WebDriverWait(self.driver, 20)


    def set_chrome_options(self):
        options = webdriver.ChromeOptions()
        # options.add_argument('--headless')
        options.add_argument('--no-sandbox')
        options.add_argument('--disable-extensions')
        options.add_argument('--disable-gpu')
        options.add_argument('--disable-web-security')
        options.add_argument('--start-maximized')
        options.add_argument('--disable-dev-shm-usage')
        options.add_argument("--log-level=3")
        options.add_experimental_option('excludeSwitches', ['enable-logging'])
        return options


    def set_webdriver(self):
        options = self.set_chrome_options()
        try:
            self.driver = webdriver.Chrome(options=options)
            self.logger.info('WebDriver started successfully.')
        except Exception as e:
            self.logger.error(f'ERR0R set_webdriver() | Failed to start WebDriver: {e}')

    def open_url(self, url: str):
        if self.driver:
            try:
                self.driver.get(url)
                self.logger.info(f'Opened URL: {url}')
            except Exception as e:
                self.logger.error(f'ERROR open_url() | Failed to open URL {url}: {e}')
        else:
            self.logger.error('ERR0R open_url() | WebDriver not initialized. You must call set_webdriver() first.')

    # def search(self, url: str, search_query: str):
    #     """
    #     Perform a search on the given URL using the provided query string.
    #     """
    #     try:
    #         search_url = url + urllib.parse.quote(search_query)
    #         self.logger.info(f'Search query submitted: {search_query}.')
    #         self.open_url(search_url)
    #     except Exception as e:
    #         self.logger.error(f'ERROR search() | Could not find element: {e}')


    def scroll_down(self, element_selector, pause_time=3, max_scrolls=2):
        """
        Rola a página para baixo várias vezes para carregar todos os itens.
        """
        last_height = self.driver.execute_script("return document.body.scrollHeight")
        
        for i in range(max_scrolls):
            time.sleep(1)  
            self.driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
            time.sleep(2)  

            new_height = self.driver.execute_script("return document.body.scrollHeight")
            # print(i, ':', len(self.driver.find_elements(By.CSS_SELECTOR, element_selector)))

            try:
                WebDriverWait(self.driver, 10).until(
                    lambda d: len(d.find_elements(By.CSS_SELECTOR, element_selector)) >= 10*(i+2)
                )
                # print(len(self.driver.find_elements(By.CSS_SELECTOR, element_selector)))
            except:
                self.logger.info("No more content to load.")

            if new_height == last_height:  # It means there are no more new content
                break
            last_height = new_height

        self.logger.info("All content loaded.")


    def get_element_list(self, element_selector):
        """
        Gets the elements from the web page. The src parameter allows the function to treat img tags.
        The function captures all the news blocks and checks if they have the according element in it. 
        If not, it registers an empty string so the order of the information is correctly organized in
        the final Excel file.        
        """
        try: 
            self.wait.until(EC.presence_of_all_elements_located((By.CSS_SELECTOR, element_selector)))
            
            products = self.driver.find_elements(By.CSS_SELECTOR, element_selector)
            code_list = []
            for product in products:
                try:
                    code = product.find_element(By.CSS_SELECTOR, 'h3 a').text
                    code_list.append(code)
                except Exception as e:
                    self.logger.warning("Error capturing a product:", e)
            self.logger.info("Finished getting elements.")
            return code_list
        except Exception as e:
            self.logger.warning("Could not get any of the products.")


    def find_element(self, element_xpath):
        try:
            return self.driver.find_element(By.XPATH, element_xpath)
        except:
            return None
    
    def span_text(self, text):
        return f"//span[text()='{text}']/following-sibling::span[@class='value']"

    def get_specs(self, product_code):
        self.open_url(f'https://www.baldor.com/catalog/{product_code}#tab="specs"')
        hp_0 = self.find_element(self.span_text('Output @ Frequency')).text
        voltage_0 = self.find_element(self.span_text('Voltage @ Frequency')).text
        rpm_0 = self.find_element(self.span_text('Synchronous Speed @ Frequency')).text
        frame = self.find_element(self.span_text('Frame')).text

        hp = hp_0.split('.')[0]
        voltage = '/'.join([v.split()[0].replace('.0', '') for v in voltage_0.split('\n')])
        rpm = rpm_0.split('RPM')[0].strip()

        return hp, voltage, rpm, frame

    def get_details(self, product_code):
        product_dict = {
            "product_id": product_code,
            "name": None,
            "description": None,
            "specs": {"hp": None, "voltage": None, "rpm": None, "frame": None},
            "bom": [],
            "assets": {"manual": None, "cad": None, "image": None}
        }

        specs = self.get_specs(product_code)
        for key, spec in zip(product_dict['specs'].keys(), specs):
            product_dict['specs'][key] = spec

        self.open_url(f'https://www.baldor.com/catalog/{product_code}#tab="parts"')
        bom_rows = self.driver.find_elements(By.CSS_SELECTOR, 'table.data-table tbody tr')
        for row in bom_rows:
            cells = row.find_elements(By.TAG_NAME, 'td')
            part_number = cells[0].text
            description = cells[1].text
            quantity = cells[2].text

            if any([part_number, description, quantity]):
                quantity = int(quantity.split('.')[0])
                product_dict['bom'].append({
                    "part_number": part_number,
                    "description": description,
                    "quantity": quantity
                })

        self.open_url(f'https://www.baldor.com/catalog/{product_code}#tab="drawings"')
        self.scroll_down(None, max_scrolls=1)
        radio_button_2d = self.driver.find_element(By.XPATH, "//input[@value='2D']")
        radio_button_2d.click()
        # dropdown = self.driver.find_element(By.XPATH, "//select[@kendo-drop-down-list]")
        # dropdown.click()
        # dwg = self.driver.find_element(By.XPATH, "//option[contains(@value, 'DWG')]")
        # dwg.click()

        # # CLICAR NO DOWNLOAD

        return product_dict



    def close_all(self):
        if self.driver:
            try:
                self.driver.quit()
                self.logger.info('WebDriver closed successfully.')
            except Exception as e:
                self.logger.error(f'ERROR close_all() | Failed to close WebDriver: {e}')
        else:
            self.logger.error('ERROR close_all() | WebDriver not initialized.')

    def to_excel(self, data):
        wb = Workbook()
        sheet = wb.active

        headers = ['Title', 'Description', 'Date', 'Filename', 'Count Query in Title and Description', 'Contains Money in Title']
        for i_col, header in enumerate(headers):
            cell_header = sheet.cell(row=1, column=i_col+1)
            cell_header.value = header
            cell_header.font = Font(bold=True, size=12)
            cell_header.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

            if data[i_col] is not None:
                for i_row, val in enumerate(data[i_col]):
                    cell_content = sheet.cell(row=i_row+2, column=i_col+1)
                    cell_content.value = val
                    sheet.row_dimensions[i_row+2].height = 60
                    if i_col > 1:  # Horizontal aligment not for columns Title and Description
                        cell_content.alignment = Alignment(horizontal='center', vertical='center')
                    else:
                        cell_content.alignment = Alignment(vertical='center', wrap_text=True)
            else:
                self.logger.error(f'ERROR to_excel() | {header} list is None.')

        sheet.column_dimensions['A'].width = 25
        sheet.column_dimensions['B'].width = 50
        for col in ['C', 'D', 'E', 'F']:
            sheet.column_dimensions[col].width = 25
        sheet.row_dimensions[1].height = 35

        sheet.title = 'NEWS'
        self.logger.info('Excel done.')
        return wb


def main():
    logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

    scraper = OEMCatalogScraper()
    scraper.set_webdriver()

    catalog_number = 110
    url = f'https://www.baldor.com/catalog#category={catalog_number}'
    # scraper.open_url(url)

    element_selector = "div[ng-repeat='product in products.matches track by product.code']"
    # scraper.scroll_down(element_selector)
    # product_codes = scraper.get_element_list(element_selector)

    # for product_code in product_codes:
    product_code = 'CEBM3546T'
    # # product_url = f'https://www.baldor.com/catalog/{product_code}'
    # # scraper.open_url(product_url)

    product_dict = scraper.get_details(product_code)
    print(product_dict)

    with open(f"{product_code}.json", "w", encoding="utf-8") as f:
        json.dump(product_dict, f, ensure_ascii=False, indent=4)

# ----- fim do for loop
    time.sleep(1)
    # scraper.close_all()

    scraper.logger.info('-'*60)


if __name__ == '__main__':
    main()
